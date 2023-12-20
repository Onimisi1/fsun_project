import os
import io
import tempfile
import pandas as pd
import openpyxl
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.http import require_POST
import uuid
from .forms import FsunForm
from .models import FieldStaffInformation
from django.core.mail import send_mail, EmailMessage

# Create your views here.


def index(request):
    form = FsunForm()
    
    if request.method == 'POST':
        form = FsunForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES.get('excel_file')
            user_email = form.cleaned_data['user_email']

            # base_name, extension = os.path.splitext(excel_file.name)
            # new_file_name = f'{base_name}_modified{extension}'
            # new_file_path = os.path.join('media', new_file_name)

            try:

                df = pd.read_excel(excel_file)

                # excel_buffer = io.BytesIO()
                # with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                #     df.to_excel(writer, index=False)


                # result = copy_excel_settings(excel_file, excel_buffer)
        
                # if result:
                    # If an error occurred in copy_excel_settings, return the error message
                    # return JsonResponse({'status': "error", 'message': "An error occurred. Please check that you are using the correct format"})


                df['Cleaned_Name'] = df['Name'].str.strip().str.lower()

                existing_data = get_existing_data(df)

                # If data exists, assign existing FSUN codes
                if not existing_data.empty:
               
                    existing_data['Cleaned_name'] = existing_data['name'].str.strip().str.lower()


                    # Merge the DataFrames based on the 'Name' column
                    df = pd.merge(df, existing_data, left_on='Cleaned_Name', right_on='Cleaned_name', how='left')


                    # Fill the 'FSUN' column with the existing FSUN codes where available
                    df['FSUN'] = df['fsun_code'].fillna(df['FSUN']).astype(str)

                    # Drop unnecessary columns
                    df = df.drop(['Cleaned_name','name', 'fsun_code'], axis=1)


                null_fsun_rows = df['FSUN'].apply(lambda x: pd.isna(x) or (isinstance(x, str) and x.strip().lower() == 'nan'))


                for index, row in df[null_fsun_rows].iterrows():
                    print(f'Savingggggg')

                    print('Before generating new FSUN codes')

                    new_fsun = generate_fsun_code()

                    df.loc[index, 'FSUN'] = new_fsun
                    
                    field_staff_instance = FieldStaffInformation(
                        fsun_code = new_fsun,
                        name = row['Cleaned_Name'],
                        state = row['State'],
                        phone_number = row['Telephone Number'],
                        address = row['Address'],
                        email = row['Email'],
                        region = row['Region'],
                        project = row['Project'],
                        role = row['Role'],
                    )
                    field_staff_instance.save()

                df = df.drop(['Cleaned_Name'], axis=1)


                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False)

            
                
                send_email_with_attachment(excel_buffer, user_email)

                return JsonResponse({'status': "success", 'message': "File Uploaded Successfully and has been sent for Approval."})
                    
            
            except Exception as e:
                # Handle exceptions (log, show error message, etc.)
                print(f"An error occurred: {e}")

                # import traceback
                # traceback_str = traceback.format_exc()

                # Log the exception traceback for debugging
                # print(f"An error occurred: {e}\n{traceback_str}")

                return JsonResponse({'status': "error", 'message': "An Error Occurred while uploading. Check that you are using the correct format"})
                           
    context = {'form': form}
    return render(request, 'index.html', context)


def send_email_with_attachment(attachment_path, user_email):
    subject = 'FSUN Approval'
    message = 'Dear Sir,' + '\n' * 2 + 'Kindly Approve'
    from_email = 'onimisiazeez7@gmail.com'
    recipient_list = [from_email]
    cc_list = [user_email]


    email = EmailMessage(subject, message, from_email, recipient_list, cc=cc_list)
                    
    # Attach the new Excel file to the email
    # email.attach_file(attachment_path)

    email.attach(f'file.xlsx', attachment_path.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Send the email
    email.send()

    # os.remove(temp_file_path)


def get_existing_data(df):

    # Extract unique names from the DataFrame for querying the database
    unique_names = df['Cleaned_Name'].unique()

    # Query the database to check for existing data and FSUN codes
    existing_data_qs = FieldStaffInformation.objects.filter(name__in=unique_names).values('name', 'fsun_code')

    # Convert the queryset to a DataFrame
    existing_data = pd.DataFrame(existing_data_qs)

    return existing_data


def generate_fsun_code():
    suffix = uuid.uuid4()
    return 'FSUN_' + str(suffix)[:8]


def copy_excel_settings(original_file, new_file):
    try:
        original_workbook = openpyxl.load_workbook(original_file)
        # new_workbook = openpyxl.load_workbook(new_file)

        new_workbook = openpyxl.load_workbook(io.BytesIO(new_file.getvalue()))
    
        for sheet_name in original_workbook.sheetnames:
            print(f'{sheet_name} in the loop')
            original_sheet = original_workbook[sheet_name]
            new_sheet = new_workbook[sheet_name]

            # Copy protection settings
            new_sheet.protection.sheet = original_sheet.protection.sheet

        # Save the changes to the new file
        new_file.seek(0)

        new_workbook.save(new_file)

    except Exception as e:
        # Handle exceptions (log, show error message, etc.)
        # raise Exception(f"An error occurred. Please check that you are using the correct format: {e}")
        return JsonResponse({'status': "error", 'message': f"An error occurred. Please check that you are using the correct format"})
